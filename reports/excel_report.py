import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

from database.db import SessionLocal
from database.models import TrackedProduct, PriceHistory


def generate_report(telegram_id: int | None = None):
    db = SessionLocal()
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    row_fill_1 = PatternFill("solid", fgColor="D6E4F0")
    row_fill_2 = PatternFill("solid", fgColor="FFFFFF")

    # ===== SHEET 1: Tracked Products =====
    ws1 = wb.active
    ws1.title = "Tracked Products"

    headers = headers = ["ID", "Query", "Active", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    query = db.query(TrackedProduct).filter(TrackedProduct.is_active == True)
    if telegram_id is not None:
        query = query.filter(TrackedProduct.telegram_id == telegram_id)
    products = query.all()
    for row, p in enumerate(products, 2):
        fill = row_fill_1 if row % 2 == 0 else row_fill_2
        data = [p.id, p.query, "Yes" if p.is_active else "No",
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col, width in enumerate([6, 15, 25, 8, 18], 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

    # ===== SHEET 2: Price History =====
    ws2 = wb.create_sheet("Price History")

    headers2 = ["ID", "Product Query", "Source", "Title", "Price ($)",
                "Rating", "Reviews", "URL", "Fetched At"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    history_query = (
        db.query(PriceHistory, TrackedProduct.query)
        .join(TrackedProduct, PriceHistory.tracked_product_id == TrackedProduct.id)
        .filter(TrackedProduct.is_active == True)
        .order_by(PriceHistory.fetched_at.desc())
    )
    if telegram_id is not None:
        history_query = history_query.filter(TrackedProduct.telegram_id == telegram_id)
    history = history_query.all()

    for row, (h, query) in enumerate(history, 2):
        fill = row_fill_1 if row % 2 == 0 else row_fill_2
        short_title = (h.title[:80] + "...") if h.title and len(h.title) > 80 else h.title
        data = [h.id, query, h.source, short_title, h.price, h.rating, h.reviews,
                h.url, h.fetched_at.strftime("%Y-%m-%d %H:%M") if h.fetched_at else ""]
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col, width in enumerate([6, 20, 16, 40, 12, 8, 10, 50, 18], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width

    # ===== SHEET 3: Price Trend Chart (for first tracked product) =====
    ws3 = wb.create_sheet("Price Trend")

    if products:
        first_product = products[0]
        trend_data = (
            db.query(PriceHistory)
            .filter(PriceHistory.tracked_product_id == first_product.id)
            .order_by(PriceHistory.fetched_at)
            .all()
        )

        ws3.cell(row=1, column=1, value="Fetched At").font = Font(bold=True)
        ws3.cell(row=1, column=2, value="Price ($)").font = Font(bold=True)

        for row, h in enumerate(trend_data, 2):
            ws3.cell(row=row, column=1, value=h.fetched_at.strftime("%Y-%m-%d %H:%M") if h.fetched_at else "")
            ws3.cell(row=row, column=2, value=h.price)

        if len(trend_data) > 1:
            chart = LineChart()
            chart.title = f"Price Trend: {first_product.query}"
            chart.y_axis.title = "Price ($)"
            chart.x_axis.title = "Date"
            chart.width = 25
            chart.height = 15

            data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(trend_data) + 1)
            cats = Reference(ws3, min_col=1, min_row=2, max_row=len(trend_data) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws3.add_chart(chart, "D2")

    os.makedirs("output", exist_ok=True)
    filename = f"output/price_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    db.close()
    print(f"✅ Report saved: {filename}")
    return filename


if __name__ == "__main__":
    generate_report()